import time
from datetime import datetime
import matplotlib.pyplot as plt
import praw
from openpyxl import Workbook, load_workbook


def top_day(data):
    wb_obj = load_workbook("Posts" + data)
    sheet_obj = wb_obj.active
    subreddits = []
    s_score = []
    all_row = list(sheet_obj.iter_rows())
    for i in range(2, sheet_obj.max_row):
        if sheet_obj.cell(row=i, column=2).internal_value not in subreddits:
            cell_val = sheet_obj.cell(row=i, column=2).internal_value
            subreddits.append(cell_val)
            subreddit_posts = list(filter(lambda x: x[1].internal_value == cell_val, all_row))
            scor_sum = 0
            for j in subreddit_posts:
                scor_sum = scor_sum + j[4].internal_value
            s_score.append((cell_val, scor_sum))

    return sorted(s_score, key=lambda x: x[1], reverse=True)


all = dict()
all["27April.xlsx"] = top_day("27April.xlsx")[:10]
all["29April.xlsx"] = top_day("29April.xlsx")[:10]
all["30April.xlsx"] = top_day("30April.xlsx")[:10]
all["07May.xlsx"] = top_day("07May.xlsx")[:10]
all["08May.xlsx"] = top_day("08May.xlsx")[:10]
all["10May.xlsx"] = top_day("10May.xlsx")[:10]
all["12May.xlsx"] = top_day("12May.xlsx")[:10]

all_reddit = [j[0] for i in all.values() for j in i]
reddits_top_10_days = dict()
for i in all_reddit:
    reddits_top_10_days[i] = all_reddit.count(i)

x = reddits_top_10_days.keys()
y = reddits_top_10_days.values()
plt.figure(figsize=(20, 15))
plt.bar(x, y,align='edge', width = 0.5, color = ['red', 'green'])
plt.xlabel('Subreddit')
plt.ylabel('Days in top 10')
plt.title('Top 10 days')
plt.show()
