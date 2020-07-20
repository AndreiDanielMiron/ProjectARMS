import time
from datetime import datetime

import praw
from openpyxl import Workbook, load_workbook

reddit = praw.Reddit(
                     user_agent="Crawler")

list_of_post_id = []
counterp = 2


def create_excel_file():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Posts"
    ws1['A1'] = 'ID'
    ws1['B1'] = 'Subreddit'
    ws1['C1'] = 'Title'
    ws1['D1'] = 'Post text'
    ws1['E1'] = 'Score'
    ws1['F1'] = 'Ration'

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Comments"
    ws2['A1'] = 'ID'
    ws2['B1'] = 'Subreddit'
    ws2['C1'] = 'Comment'

    data = datetime.now().strftime("%d%B")
    wb.save(filename='Posts' + data + '.xlsx')
    wb2.save(filename='Comments' + data + '.xlsx')


def split_list(lista, size):
    new_list = []
    position = 0
    if size > len(lista):
        new_list.append(lista)
    else:
        while position + size < len(lista):
            new_list.append(lista[position: position + size])
            position += size
        else:
            new_list.append(lista[position:])
    return new_list


# def get_top_posts(subreddit, time_period='day'):
#     lastday_subreddit = list(reddit.subreddit(subreddit).top(time_period))
#     global counterp
#
#     for lista in split_list(lastday_subreddit, 20):
#         start_time = time.time()
#         for submission in lista:
#             ws1['A' + str(counterp)] = submission.id
#             ws1['B' + str(counterp)] = subreddit
#             ws1['C' + str(counterp)] = submission.title
#             ws1['D' + str(counterp)] = submission.selftext
#             ws1['E' + str(counterp)] = submission.score
#             ws1['F' + str(counterp)] = submission.upvote_ratio
#             list_of_post_id.append((subreddit, submission.id))
#             counterp += 1
#         end_time = time.time()
#         print(end_time - start_time)
#         time_spent = end_time - start_time
#         if time_spent < 60:
#             time.sleep(60 - time_spent)
#
#
# def populate_today_top(list_of_subreddits):
#     for subreddit in list_of_subreddits:
#         try:
#             print(50 * "*")
#             print(subreddit.title)
#             get_top_posts(subreddit.display_name)
#         except Exception as e:
#             print(e)
#     wb.save(filename='Posts' + data + '.xlsx')


def get_top_comments(data):
    counter = 2
    # data = datetime.now().strftime("%d%B")
    wb = load_workbook('Comments' + data)
    ws1 = wb.active
    contor = 0

    for post in list_of_post_id:
        try:
            start_time = time.time()
            submission = reddit.submission(post)
            submission.comments.replace_more(limit=None)
            for top_level_comment in submission.comments:
                ws1['A' + str(counter)] = post
                ws1['B' + str(counter)] = top_level_comment.subreddit.display_name
                ws1['C' + str(counter)] = top_level_comment.body
                counter += 1
                for second_level_comment in top_level_comment.replies:
                    ws1['A' + str(counter)] = post
                    ws1['B' + str(counter)] = second_level_comment.subreddit.display_name
                    ws1['C' + str(counter)] = second_level_comment.body
                    counter += 1
            end_time = time.time()
            time_spent = end_time - start_time
            # if time_spent < 60:
            print(time_spent)
            contor = contor + 1
            if contor == 19:
                contor = 0
                time.sleep(60)
        except Exception as e:
            print(e)
            print("Err Aici 1------>>>>>")

    wb.save(filename='Comments' + data)


# def start_data_extraction():
#     popular_subreddits = list(reddit.subreddits.popular())
#     populate_today_top(popular_subreddits)
# get_top_comments()


def get_comm(data: str):
    wb_obj = load_workbook('Posts' + data)
    current = 2
    sheet_obj = wb_obj.active
    global list_of_post_id
    for q in range(1, sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=current, column=1)
        current = current + 1
        list_of_post_id.append(cell_obj.value)

    get_top_comments(data)
    list_of_post_id = []


# create_excel_file()
# data = datetime.now().strftime("%d%B")
# wb = load_workbook('Posts' + data + '.xlsx')
# ws1 = wb.active
# start_data_extraction()

# get_comm("27April.xlsx")
# get_comm("29April.xlsx")
# get_comm("30April.xlsx")
get_comm("07May.xlsx")
get_comm("08May.xlsx")
# get_comm("10May.xlsx")
# get_comm("11May.xlsx")
# get_comm("12May.xlsx")
